#read_excel.py
import openpyxl

url = "datas/"

lang_map = {"rus": 1, "kaz": 2, "eng": 3}

def read_excel_cell(file_name, sheet_name, lang, give_info):
    # Словари для преобразования языка и типа информации в индексы
    lang_map = {"rus": 1, "kaz": 2, "eng": 3}

    info_map = {"pro_act": 1,
                "admission": 2,
                "score": 3,
                "contact": 4,
                "name": 5}

    # Получение индексов строки и столбца
    row = lang_map.get(lang)
    col = info_map.get(give_info)

    # Открытие файла Excel
    workbook = openpyxl.load_workbook(f"{url}{file_name}.xlsx", read_only=True)
    sheet = workbook[sheet_name]

    # Чтение значения ячейки
    return sheet.cell(row=row, column=col).value

def get_name_and_index_and_info(file_name,lang):
    workbook = openpyxl.load_workbook(f"{url}{file_name}.xlsx", read_only=True)
    sheets = workbook.sheetnames[:-1]
    info = workbook['info'].cell(row=lang_map[lang], column=1).value
    name_and_index = []
    for sheet_name in sheets:
        sheet = workbook[sheet_name]
        name_and_index.append([sheet.cell(row=lang_map[lang], column=5).value,sheet_name])
    return (name_and_index,info)

def get_profession_index(file_names):
    result = set()
    for file_name in file_names:
        result = result | set(openpyxl.load_workbook(f"{url}{file_name}.xlsx", read_only=True).sheetnames[:-1])
    return result

def get_text_answer(sheet_name,lang):
    sheet = openpyxl.load_workbook(f"{url}standard_answer.xlsx", read_only=True)[sheet_name]
    return sheet.cell(row=lang_map[lang], column=1).value

def get_text_for_info(file_name,sheet_name,lang):
    sheet = openpyxl.load_workbook(f"{url}{file_name}.xlsx", read_only=True)[sheet_name]
    return sheet.cell(row=lang_map[lang], column=6).value